import os
import webview
import requests
import openpyxl
import pandas as pd
import concurrent.futures
from PyPDF2 import PdfReader
from bs4 import BeautifulSoup
from flair.data import Sentence
from urllib.parse import urlparse
from flair.models import SequenceTagger
from openpyxl.utils import get_column_letter
import time
from llama_index.llms.huggingface import HuggingFaceInferenceAPI
from llama_index.core import PromptTemplate
HF_token = '******'

llm = HuggingFaceInferenceAPI(model_name="mistralai/Mixtral-8x7B-Instruct-v0.1", token=HF_token)


tagger = SequenceTagger.load('ner')
def extract_person_names(text,company,url):
    try:
        template = """
        Question: {question}
        Context: {context_str}
        Answer: Let's think step by step.
        """
        prompt_template = PromptTemplate(template=template, input_variables="question, context_str")

        question = f"""Make a list of the names of people who are being interviewed
          or are giving testimonials of their experience and how they got in the company {company}. 
          Do not say anything else. Keep in mind that there may not be a single one that have given an 
          interview- if that it is the case, do not return any names. Only those people (if there is any) who 
          are of {company} and are telling about their experience in that job, and how they applied, prepared, etc.
          Here's an example:
          'John Satimburwa is a finance graduate, working for Laing O’Rourke in our head office in Dartford. He started the graduate development programme in 2023. Read about his experience with Laing O'Rourke so far.
            When did you join Laing O'Rourke?
          I joined the graduate development programme in 2023.

Give us a brief overview of your day to day responsibilities. 

No two days are the same with this role. One day I am having meeting with business partners about their budgets and the next day I am processing prepayments and accruals and posting them into the finance system that we use. In summary I am getting full exposure to the different facets of the finance function at Laing O'Rourke and the work I do enables those above me to do theirs.
What drew you to apply for your role at LOR? 

What stood out to me about LOR was the passion that the company has for the work that they do. In addition LOR is known as a company at the forefront of innovation in the sector and I wanted to work for a company that was pushing the boundaries and daring to be great. That gave me great excitement as a graduate and made me want to be somewhere where I can contribute to such a cause.
Do you have any hints and tips for the recruitment process? 

Just be yourself. I know it sounds cliché but it really is true. If you pretend to be someone who you’re not then you will sell yourself short and end up regretting it. In addition make sure you are prepared. The application process is demanding so be ready to meet those demands. The best way to do that is to do the appropriate research and preparation for each stage.

What advice would you give to undergrads looking for a graduate role? 

Apply for as many roles as you can. This is tiring and can be deflating when those rejections come in, but the more roles you apply for the greater a chance you give yourself to land a role. I had a schedule for applying for jobs where I set aside one hour aside three days a week. Furthermore, I applied for roles outside the "accounting and finance description" even though a job within finance was my preference. This is because I wanted to give myself as many options as possible, which actually allowed me to come across websites where certain jobs where being advertised that I hadn't seen anywhere else (LOR being one of them).
' For this, you would have to answer John Satimburwa"""

        response = llm.complete(prompt_template.format(question=question, context_str=text))
        response_text = response.text
        sentence = Sentence(response_text)
        tagger.predict(sentence)
        person_names = set(entity.text for entity in sentence.get_spans('ner') if entity.tag == 'PER')
        print(f'Done for company {company} and webpage {url}')
        ans =  list(person_names)

        ans = [s.replace(company, "") for s in ans]
        return ans
    except Exception as e:
        print(e)

        return []

def extract_text(web_url, company):
    try:
        if web_url.lower().endswith('.pdf'):
            pdf_filename = os.path.basename(urlparse(web_url).path)

            if not os.path.exists('documents'):
                os.makedirs('documents')

            pdf_filepath = os.path.join('documents', pdf_filename)

            response = requests.get(web_url)
            with open(pdf_filepath, 'wb') as f:
                f.write(response.content)

            with open(pdf_filepath, 'rb') as f:
                reader = PdfReader(f)
                text = ""
                for page in reader.pages:
                    text += page.extract_text()

            print(f'Done for {web_url}')
            return text

        else:
            response = requests.get(web_url)
            soup = BeautifulSoup(response.content, 'html.parser')
            text = soup.get_text()
            print(f'Done for {web_url}')
            return text

    except Exception as e:
        print(f'Failed at {web_url} and {company}: {str(e)}')




def google_search(query,num):
    query = query.replace(' ', '+')
    url = f"https://www.google.com/search?q={query}?num={num}"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.121 Safari/537.36"
    }
    response = requests.get(url, headers=headers)
    response.raise_for_status()

    soup = BeautifulSoup(response.text, 'html.parser')

    links = []
    for item in soup.find_all('div', class_='yuRUbf'):
        a_tag = item.find('a')
        if a_tag and a_tag.get('href'):
            links.append(a_tag['href'])

    return links
def fetch_details(company_name):
    try:
        if company_name==None:
            return 'Please Enter Company Name'

        company = company_name
        queries = [
            f'"{company}" "employee testimonials" "UK" -jobs -recruitment -hiring',
            f'"{company}" "employee experiences" "UK" -jobs -recruitment -hiring',
            f'"working at {company}" "my experience" "UK" -jobs -recruitment -hiring',
            f'"life at {company}" "career journey" "UK" -jobs -recruitment -hiring',
            f'"{company}" "employee spotlight" "UK" -jobs -recruitment -hiring',
            f'"{company}" "staff interviews" "UK" -jobs -recruitment -hiring',
            f'"{company}" "meet our team" "UK" -jobs -recruitment -hiring',
            f'"{company}" "employee profiles" "UK" -jobs -recruitment -hiring',
            f'"{company}" "career stories" "UK" -jobs -recruitment -hiring',
            f'site:linkedin.com/in/ "worked at {company}" "UK"',
            f'site:youtube.com "{company}" "employee interview" "UK"',
        ]
        company_testimonials = []
        for query in queries:
            try:
                results = google_search(query, 10)
                company_testimonials.extend(results)
            except Exception as e:
                print(f"Error with query '{query}': {e}")

        company_testimonials = list(set(company_testimonials))
        company_testimonials = [{
            'company':company,
            'web_url': company_testimonials
        }]
        def process_testimonial(comp):
            company = comp['company']
            web_sites = []
            for url in comp['web_url']:
                text = extract_text(url, company)

                if text:
                    web_sites.append({'company': company, 'web_content': str(text),'url':url})
            return web_sites

        all_web_sites = []

        with concurrent.futures.ThreadPoolExecutor() as executor:
            results = executor.map(process_testimonial, company_testimonials)

            for result in results:
                all_web_sites.extend(result)


        row_list = []

        def process_website(part):
            try:
                text = part['web_content']
                company = part['company']
                url = part['url']
                names = extract_person_names(text,company,url)

                return [{'Name': name, 'Company': company,'Testimonial Page':url} for name in names]
            except:
                print(names)
                return []

        for web_site in all_web_sites:
            result = process_website(web_site)
            row_list.extend(result)


        file_name = f'{company_name}_list.xlsx'
        df = pd.DataFrame(row_list)
        df['Testimonial Page'] = ''
        df.to_excel(file_name,index=False)


        update_file(file_name)
        return f'Successfully done for {file_name}'
    except Exception as e:
        print(e)
        print('Something went wrong!')


def update_file(file_name):
    if not os.path.isfile(file_name):
        print(f"File with the name '{file_name}' not found.")
        return 'File Name not found!'
    df = pd.read_excel(file_name)

    if not 'Linkedin Profile' in df.columns:
        df['Linkedin Profile'] = ''

    for index, row in df.iterrows():
        if 'Link' not in str(row['Linkedin Profile']):
            try:
                name = row['Name']
                company_name = row['Company']
                query = f'{name} UK {company_name} profile linkedin'
                link = google_search(query, 10)[0]
                df.at[index, 'Linkedin Profile'] = link
                print(f'Done for index {index}')
            except Exception as e:

                print("Max attempts reached")
                print(e)
                df.at[index, 'Linkedin Profile'] = 'IP BLOCKED'
                pass
        if 'Link' not  in str(row['Testimonial Page']):
            try:
                name = row['Name']
                company_name = row['Company']
                query = f'{name} UK {company_name} testimonail client experience'
                link = google_search(query, 10)[0]
                df.at[index, 'Testimonial Page'] = link
                print(f'Done for testimonial index {index}')
            except Exception as e:

                print("Max attempts reached")
                print(e)
                df.at[index, 'Testimonial Page'] = 'IP BLOCKED'
                pass



    df.to_excel(file_name, index=False)

    def convert_urls_to_hyperlinks(excel_file, sheet_name, url_column):
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]

        col_letter = get_column_letter(url_column)

        for row in range(2, sheet.max_row + 1):
            cell = sheet[f"{col_letter}{row}"]
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                sheet[f"{col_letter}{row}"].hyperlink = cell.value
                sheet[f"{col_letter}{row}"].value = 'Link'
                sheet[f"{col_letter}{row}"].style = "Hyperlink"

        workbook.save(excel_file)

    excel_file = file_name
    sheet_name = "Sheet1"
    url_column = 3

    convert_urls_to_hyperlinks(excel_file, sheet_name, url_column)
    url_column = 4
    convert_urls_to_hyperlinks(excel_file, sheet_name, url_column)
    return f'Successfully updated data to {excel_file}'
class API:
    def check_data(self, name):
        return fetch_details(name)
    def update_data(self,name):
        return update_file(name)

html_content = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Scrape Web</title>
    <style>
        body { font-family: Arial, sans-serif; }
        .container { text-align: center; margin-top: 50px; }
        .timer { font-size: 14px; margin-top: 20px; }
        .result { font-size: 16px; color: green; margin-top: 20px; }
    </style>
</head>
<body>
    <div class="container">
        <h4>Enter Company Name</h4>
        <input type="text" id="nameInput" placeholder="Enter name">
        <button onclick="startProcess()">Start Scraping</button>

        <h4>Enter File Name</h4>
        <input type="text" id="fileNameInput" placeholder="Enter File Name">

        <button onclick="startUpdate()">Update File</button>
        <div class="timer">
            <p><span id="timerDisplay">0m 0.0s</span> (expected 15-20m)</p>
        </div>
        <div class="result" id="resultDisplay">
            <!-- The result will be displayed here -->
        </div>
    </div>

    <script>
        let timer;
        let startTime;

        async function startProcess() {
            startTimer();
            const name = document.getElementById('nameInput').value;
            const result = await pywebview.api.check_data(name);
            stopTimer();
            document.getElementById('resultDisplay').textContent = result;
        }

        async function startUpdate() {
            startTimer();
            const name = document.getElementById('fileNameInput').value;
            const result = await pywebview.api.update_data(name);
            stopTimer();
            document.getElementById('resultDisplay').textContent = result;
        }

        function startTimer() {
            startTime = Date.now();
            timer = setInterval(updateTimer, 100);
        }

        function updateTimer() {
            let elapsed = (Date.now() - startTime) / 1000; // elapsed time in seconds
            let minutes = Math.floor(elapsed / 60);
            let seconds = (elapsed % 60).toFixed(1);
            document.getElementById('timerDisplay').textContent = `${minutes}m ${seconds}s elapsed`;
        }

        function stopTimer() {
            clearInterval(timer);
        }
    </script>
</body>
</html>
"""

def start_gui():
    api = API()
    webview.create_window('Web Scraper', html=html_content, width=400, height=300, js_api=api)
    webview.start(gui='winforms')  # Use the CEF (Chromium) renderer for better compatibility

if __name__ == '__main__':
    start_gui()
