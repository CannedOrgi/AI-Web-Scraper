import sys
from PyQt6.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QLineEdit
from PyQt6.QtCore import QTimer, QElapsedTimer

import os
import webview
import requests
import openpyxl
import pandas as pd
import concurrent.futures
from PyPDF2 import PdfReader
from bs4 import BeautifulSoup
from flair.data import Sentence
from urllib.parse import urlparse
from flair.models import SequenceTagger
from openpyxl.utils import get_column_letter


tagger = SequenceTagger.load('ner')
def extract_person_names(text,company,url):
    try:
        sentence = Sentence(text)
        tagger.predict(sentence)
        person_names = set(entity.text for entity in sentence.get_spans('ner') if entity.tag == 'PER')
        print(f'Done for company {company} and webpage {url}')
        ans =  list(person_names)

        ans = [s.replace(company, "") for s in ans]
        return ans
    except Exception as e:
        print(e)

        return []

def extract_text(web_url, company):
    try:
        if web_url.lower().endswith('.pdf'):
            pdf_filename = os.path.basename(urlparse(web_url).path)
            
            if not os.path.exists('documents'):
                os.makedirs('documents')
            
            pdf_filepath = os.path.join('documents', pdf_filename)
            
            response = requests.get(web_url)
            with open(pdf_filepath, 'wb') as f:
                f.write(response.content)
            
            with open(pdf_filepath, 'rb') as f:
                reader = PdfReader(f)
                text = ""
                for page in reader.pages:
                    text += page.extract_text()
            
            print(f'Done for {web_url}')
            return text

        else:
            response = requests.get(web_url)
            soup = BeautifulSoup(response.content, 'html.parser')
            text = soup.get_text()
            print(f'Done for {web_url}')
            return text

    except Exception as e:
        print(f'Failed at {web_url} and {company}: {str(e)}')




def google_search(query,num):
    query = query.replace(' ', '+')
    url = f"https://www.google.com/search?q={query}?num={num}"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.121 Safari/537.36"
    }
    response = requests.get(url, headers=headers)
    response.raise_for_status() 

    soup = BeautifulSoup(response.text, 'html.parser')

    links = []
    for item in soup.find_all('div', class_='yuRUbf'):
        a_tag = item.find('a')
        if a_tag and a_tag.get('href'):
            links.append(a_tag['href'])

    return links
def fetch_details(company_name):
    try:
        if company_name==None:
            return 'Please Enter Company Name'

        company = company_name
        query = f'{company} UK Testimonials'
        try:
            company_testimonials = google_search(query,1)
        except:
            return 'IP BLOCKED'


        company_testimonials = [{
            'company':company,
            'web_url': company_testimonials
        }]
        def process_testimonial(comp):
            company = comp['company']
            web_sites = []
            for url in comp['web_url']:
                text = extract_text(url, company) 

                if text:
                    web_sites.append({'company': company, 'web_content': str(text),'url':url})
            return web_sites

        all_web_sites = []

        with concurrent.futures.ThreadPoolExecutor() as executor:
            results = executor.map(process_testimonial, company_testimonials)

            for result in results:
                all_web_sites.extend(result)


        row_list = []

        def process_website(part):
            try:
                text = part['web_content']
                company = part['company']
                url = part['url']
                names = extract_person_names(text,company,url)
                
                return [{'Name': name, 'Company': company,'Testimonial Page':url} for name in names]
            except:
                print(names)
                return []

        for web_site in all_web_sites:
            result = process_website(web_site)
            row_list.extend(result)       


        file_name = f'{company_name}_list.xlsx'
        df = pd.DataFrame(row_list)
        df.to_excel(file_name,index=False)


        update_file(file_name)
        return f'Successfully done for {file_name}'
    except Exception as e:
        print(e)
        return 'Something went wrong!'


def update_file(file_name):
    if not os.path.isfile(file_name):
        print(f"File with the name '{file_name}' not found.")
        return 'File Name not found!'
    df = pd.read_excel(file_name)

    if not 'Linkedin Profile' in df.columns:
        df['Linkedin Profile'] = ''

    for index,row in df.iterrows():
        if 'Link' in row['Linkedin Profile']:
            continue
        try:
            name = row['Name']
            query = f'{name} UK profile linkedin'
            link = google_search(query,10)[0]
            df.at[index,'Linkedin Profile'] = link
            print(f'Done for index {index}')
        except Exception as e:
            print(e)
            df.at[index,'Linkedin Profile'] = 'IP BLOCKED'
            pass
    df.to_excel(file_name,index=False)
  
    def convert_urls_to_hyperlinks(excel_file, sheet_name, url_column):
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]

        col_letter = get_column_letter(url_column)

        for row in range(2, sheet.max_row + 1): 
            cell = sheet[f"{col_letter}{row}"]
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                sheet[f"{col_letter}{row}"].hyperlink = cell.value
                sheet[f"{col_letter}{row}"].value = 'Link'
                sheet[f"{col_letter}{row}"].style = "Hyperlink"

        workbook.save(excel_file)

    excel_file = file_name
    sheet_name = "Sheet1"
    url_column = 3  

    convert_urls_to_hyperlinks(excel_file, sheet_name, url_column)
    url_column = 4
    convert_urls_to_hyperlinks(excel_file, sheet_name, url_column)
    return f'Succesfully updated data to {excel_file}'
import sys
from PyQt6.QtWidgets import QApplication, QWidget, QVBoxLayout, QLineEdit, QPushButton, QLabel, QHBoxLayout
from PyQt6.QtCore import QTimer, QElapsedTimer, QThread, pyqtSignal


class WorkerThread(QThread):
    result_ready = pyqtSignal(str)  # Signal to send the result back to the main thread

    def __init__(self, function, *args, **kwargs):
        super().__init__()
        self.function = function
        self.args = args
        self.kwargs = kwargs

    def run(self):
        result = self.function(*self.args, **self.kwargs)
        self.result_ready.emit(result)  # Emit the result when done


class MyApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("PyQt6 App")
        self.setGeometry(100, 100, 400, 300)

        layout = QVBoxLayout()

        self.entry1 = QLineEdit(self)
        self.entry1.setPlaceholderText("Enter Company Name")
        layout.addWidget(self.entry1)

        self.button1 = QPushButton("Start Scraping", self)
        self.button1.clicked.connect(self.on_button1_click)
        layout.addWidget(self.button1)

        self.entry2 = QLineEdit(self)
        self.entry2.setPlaceholderText("Enter File Name(with file extensions)")
        layout.addWidget(self.entry2)

        self.button2 = QPushButton("Update File", self)
        self.button2.clicked.connect(self.on_button2_click)
        layout.addWidget(self.button2)

        # Timer display label
        self.timer_label = QLabel("0m 0.0sec", self)
        self.timer_label.setStyleSheet("font-size: 14pt; color: white;")
        layout.addWidget(self.timer_label)

        self.label_result = QLabel("", self)
        self.label_result.setStyleSheet("font-size: 14pt;")
        layout.addWidget(self.label_result)

        self.label_click_count = QLabel("", self)
        self.label_click_count.setStyleSheet("font-size: 10pt; color: gray;")

        bottom_layout = QHBoxLayout()
        bottom_layout.addWidget(self.label_click_count)
        bottom_layout.addStretch()

        layout.addLayout(bottom_layout)

        self.setLayout(layout)

        # Timer setup
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_timer)

        self.elapsed_timer = QElapsedTimer()

    def on_button1_click(self):
        text1 = self.entry1.text()
        self.start_timer()
        self.worker_thread = WorkerThread(fetch_details, text1)
        self.worker_thread.result_ready.connect(self.on_worker_done)
        self.worker_thread.start()

    def on_button2_click(self):
        text2 = self.entry2.text()
        self.start_timer()
        self.worker_thread = WorkerThread(update_file, text2)
        self.worker_thread.result_ready.connect(self.on_worker_done)
        self.worker_thread.start()

    def start_timer(self):
        self.elapsed_timer.start()  # Start the elapsed timer
        self.timer.start(100)  # Update every 100 ms

    def update_timer(self):
        elapsed_time = self.elapsed_timer.elapsed()  # Time in milliseconds
        minutes = elapsed_time // 60000
        seconds = (elapsed_time % 60000) / 1000.0
        self.timer_label.setText(f"{minutes}m {seconds:.1f}sec")

    def on_worker_done(self, result):
        self.stop_timer()
        self.label_click_count.setText(result)

    def stop_timer(self):
        self.timer.stop()



if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    ex.show()
    sys.exit(app.exec())
