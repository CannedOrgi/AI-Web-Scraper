import requests
import heapq
import re
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from collections import deque
import os
import json
import openpyxl
import pandas as pd
import concurrent.futures
from PyPDF2 import PdfReader
from openpyxl.utils import get_column_letter
import time
from openai import OpenAI
import tiktoken
import company_names_list_making

companies = company_names_list_making.companies
encoder = tiktoken.get_encoding("gpt2")

YOUR_API_KEY = "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"
client = OpenAI(api_key="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", base_url="https://api.perplexity.ai")

llm_url = "https://api.perplexity.ai/chat/completions"

def google_search(query,num):
    query = query.replace(' ', '+')
    url = f"https://www.google.com/search?q={query}&num={num}"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.121 Safari/537.36"
    }
    response = requests.get(url, headers=headers)
    response.raise_for_status()

    soup = BeautifulSoup(response.text, 'html.parser')

    links = []
    for item in soup.find_all('div', class_='yuRUbf'):
        a_tag = item.find('a')
        if a_tag and a_tag.get('href'):
            links.append(a_tag['href'])

    return links

def extract_text(web_url, company):
    try:
        if web_url.lower().endswith('.pdf'):
            pdf_filename = os.path.basename(urlparse(web_url).path)

            if not os.path.exists('documents'):
                os.makedirs('documents')

            pdf_filepath = os.path.join('documents', pdf_filename)

            response = requests.get(web_url)
            with open(pdf_filepath, 'wb') as f:
                f.write(response.content)

            try:
                with open(pdf_filepath, 'rb') as f:
                    reader = PdfReader(f)
                    text = ""
                    for page in reader.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text
                print(f'Done for {web_url}')
               
                return text
            except Exception as e:
                print(f"Failed to read PDF {web_url}: {e}")
                return None  # Return None if PDF reading fails
        else:
            response = requests.get(web_url)
            soup = BeautifulSoup(response.content, 'html.parser')
            text = soup.get_text()
            print(f'Done for {web_url}')
            return text

    except Exception as e:
        print(f'Failed at {web_url} and {company}: {str(e)}')
        return None 


def extract_list_from_response(response_text):
    try:
        start_index = response_text.find('[')
        end_index = response_text.find(']')


        if start_index != -1 and end_index != -1:
        
            list_string = response_text[start_index + 1:end_index]

        
            name_list = [name.strip() for name in list_string.split(',') if name.strip()]
            return name_list
        else:
            return [] 
    except Exception as e:
        print(f"Error extracting list: {e}")
        return []
    
def extract_dict_from_string(response_text):
    try:
        start_index = response_text.find('{')
        end_index = response_text.rfind('}')
        
        if start_index != -1 and end_index != -1:
            dict_string = response_text[start_index:end_index + 1]
            
            dict_string = dict_string.replace("'", "\"")
            extracted_dict = json.loads(dict_string)
            return extracted_dict
        else:
            return {}
        
    except Exception as e:
        print(f"Error extracting dictionary: {e}")
        return {}



def get_links(url, domain):
    try:
        response = requests.get(url)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        links = set()
        
        for link in soup.find_all('a', href=True):
            href = link.get('href')
            full_url = urljoin(url, href)
            parsed_full_url = urlparse(full_url)
            netloc = parsed_full_url.netloc.replace('www.', '')
           
            if netloc == domain or netloc.endswith('.' + domain):
                links.add(full_url)
                
        return links
    except requests.RequestException as e:
        print(f"Error fetching {url}: {e}")
        return set()


def extract_person_names(text, url,company):
    messages = [
      {
          "role": "system",
          "content": (
              "Be informative"
          ),
      },
      {
          "role": "user",
          "content": (
              f"""\"{text}\"\n\ntell me if this text has
                          any person talking about how they got into the {company} / what they did to 
                          prepare for interviews, how they applied, etc. (and tell me their names) Note that there's a very slight 
                          chance of this being there, so tell me the person(s) names only if you're sure 
                          about it. Otherwise, MENTION NO NAMES AT ALL. write the name(s) in list format ([name1, name2]). Don't give me a list if there are no names"""
                     ),
      },
              ]
  
    response = client.chat.completions.create(
          model="llama-3.1-sonar-small-128k-chat",
          messages=messages,
      )
    response_text = response.choices[0].message.content
    extracted_names = extract_list_from_response(response_text)
    invalid_names = {"No names", "None", "", "N/A"}
    extracted_names = [name.strip().title() for name in extracted_names if name.strip() not in invalid_names]
    if extracted_names:
        print("Exploring...")
        print(f"{text}\n\n\n")
    #    #print(response_text)
    elif not extracted_names:
        print(".")
        print("No names extracted")
    time.sleep(0.4)
    return extracted_names

    
def explore_website(start_url, max_depth=5):
    try:
        print("Process will take about 5 mins...")
        messages = [
            {
                "role": "system",
                "content": (
                    "Be informative"
                ),
            },
            {
                "role": "user",
                "content": (
                    f"which company's url is this {start_url}. If it's a linkedin url, tell me which company the person got into instead, not the name of the company LinkedIn. reply only with the name of the company"
                ),
            },
        ]

        response = client.chat.completions.create(
            model="llama-3.1-sonar-small-128k-chat",
            messages=messages,
        )
        company = response.choices[0].message.content
        parsed_start_url = urlparse(start_url)
        domain = parsed_start_url.netloc.replace('www.', '')
        visited = set()
        queued_links = set()
        queue = []
        heapq.heappush(queue, (0, (start_url, 0)))
        queued_links.add(start_url)

        while queue and len(visited) < 150:
            _, (url, depth) = heapq.heappop(queue)

            if depth > max_depth:
                continue

            if url in visited:
                continue

            visited.add(url)
            print(f"Exploring: {url} (Depth: {depth})")

            links = get_links(url, domain)

            for link in links:
                if link not in visited and link not in queued_links:
                    # Assign priority
                    priority = 0
                    if any(pattern in link.lower() for pattern in ['uk', 'apprenticeships', 'career' 'graduate', 'testimonial', 'success', 'stories', 'early', 'program']):
                        priority = -depth  # Higher priority
                    else:
                        priority = depth
                    heapq.heappush(queue, (priority, (link, depth + 1)))
                    queued_links.add(link)

        
        company_testimonials = list(visited)
        
        company_testimonials = [{'company': company, 'web_url': company_testimonials}]

        def process_testimonial(comp):
            try:    
                company = comp['company']
                web_sites = []
                for url in comp['web_url']:

                    text = extract_text(url, company)
                    print('.')

                    if text:
                        web_sites.append({'company': company, 'web_content': str(text), 'url': url})
                return web_sites
            except Exception as e:
                print(f"Error in extracting text: {e}")
                return []

        all_web_sites = []
        with concurrent.futures.ThreadPoolExecutor() as executor:
            results = executor.map(process_testimonial, company_testimonials)
            
            print('Processing Explored Links...')
            for result in results:
                all_web_sites.extend(result)
        

        row_list = []

        def process_website(part):
            try:
                text = part['web_content']
                text = text.strip().replace("\n\n", " ")
                text = re.sub(r'\s+', ' ', text)
                num_tokens = len(encoder.encode(text))
                if num_tokens>40000:
                    print("No names extracted")
                    return []
                company = part['company']
                url = part['url']
                names = extract_person_names(text, url, company)
                print('.')

                if names:
                    time.sleep(1)
                    
                    messages = [
                        {
                            "role": "system",
                            "content": (
                                "Be informative"
                            ),
                        },
                        {
                            "role": "user",
                            "content": (
                                f"""For these {names}, find what their job role was in this: {text}.
                                Return a JSON object with the format {{\"name1\": \"job_title1\", \"name2\": \"job_title2\"}}.
                                Don't return anything else."""
                            ),
                        },
                    ]

                    response = client.chat.completions.create(
                        model="llama-3.1-sonar-small-128k-chat",
                        messages=messages,
                    )
                   
                    dict_jobs = extract_dict_from_string(response.choices[0].message.content)
                    if not isinstance(dict_jobs, dict):
                        dict_jobs = {}
                    
                if names is None:
                    names = []
                return [{'Name': name, 'Company': company, 'Testimonial Page': url, 'Job Title': dict_jobs[name]} for name in names]
            except Exception as e:
                print(f"Error processing {part['url']}: {e}")
                if 'Request rate limit exceeded' in str(e):
                    time.sleep(2)
                return []

        for web_site in all_web_sites:
            result = process_website(web_site)
            row_list.extend(result)

        file_name = f'{company}_list.xlsx'
        df = pd.DataFrame(row_list)
        if df.empty:
            print("No data to write to Excel; the DataFrame is empty.")
            return

        df.drop_duplicates(subset=['Name', 'Company'], inplace=True)

        df = df[df['Name'].notnull() & ~df['Name'].isin(["No names", "None", "", "N/A","Mention No Names At All", "No Names Mentioned"])]
        df.to_excel(file_name, index=False)
        print(update_file(file_name))

        print(f'Successfully done for {file_name}')
        return
    except Exception as e:
        print(e)
        print('Something went wrong!')


def update_file(file_name):
    if not os.path.isfile(file_name):
        print(f"File with the name '{file_name}' not found.")
        return 'File Name not found!'
    df = pd.read_excel(file_name)

    if not 'Linkedin Profile' in df.columns:
        df['Linkedin Profile'] = ''

    for index, row in df.iterrows():
        if 'Link' not in str(row['Linkedin Profile']):
            try:
                name = row['Name']
                company_name = row['Company']
                job = row['Job Title']
                query = f'{name} UK {company_name} {job} linkedin profile'
                print(query)
                link = google_search(query, 10)[0]
                df.at[index, 'Linkedin Profile'] = link
                print(f'Done for index {index}')
                time.sleep(1)
            except Exception as e:

                print("Max attempts reached for Google Search.\nPlease try again with a different network or after a few hours.")
                print(e)
                df.at[index, 'Linkedin Profile'] = 'IP BLOCKED'
                pass

        if pd.isnull(row['Testimonial Page']) or not str(row['Testimonial Page']).startswith('http'):
            try:
                name = row['Name']
                company_name = row['Company']
                query = f'{name} UK {company_name} testimonial client experience'
                link = google_search(query, 10)[0]
                df.at[index, 'Testimonial Page'] = link
                print(f'Done for testimonial index {index}')
            except Exception as e:

                print("Max attempts reached")
                print(e)
                df.at[index, 'Testimonial Page'] = 'IP BLOCKED'
                pass



    df.to_excel(file_name, index=False)

    def convert_urls_to_hyperlinks(excel_file, sheet_name, url_column):
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]

        col_letter = get_column_letter(url_column)

        for row in range(2, sheet.max_row + 1):
            cell = sheet[f"{col_letter}{row}"]
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                sheet[f"{col_letter}{row}"].hyperlink = cell.value
                sheet[f"{col_letter}{row}"].value = 'Link'
                sheet[f"{col_letter}{row}"].style = "Hyperlink"

        workbook.save(excel_file)

    excel_file = file_name
    sheet_name = "Sheet1"
    url_column = 3

    convert_urls_to_hyperlinks(excel_file, sheet_name, url_column)
    url_column = 5
    convert_urls_to_hyperlinks(excel_file, sheet_name, url_column)
    return f'Successfully updated data to {excel_file}'


    

    
if __name__ == "__main__":
    for i in range(len(companies)):
            if companies[i]=='https://www.next15.com/':
                for j in range(i, len(companies)):


                    explore_website(companies[j])
                break

    print(f"Web scraping completed.")
